import openpyxl
import random
import os

# --- YOUR COLORS ---
# We map the Hex codes (without #) to Difficulty
COLOR_MAP = {
    "D9EAD3": "Easy",    # Light Green
    "B6D7A8": "Medium",  # Medium Green
    "93C47D": "Hard"     # Dark Green
}

class DSALoader:
    def __init__(self, excel_path, sheet_name):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.questions_cache = []


    def _normalize_hex(self, hex_code):
        """
        Excel often returns 'FFD9EAD3' (ARGB). We want 'D9EAD3' (RGB).
        """
        if not hex_code or hex_code == '00000000': 
            return None
        
        hex_str = str(hex_code).upper()
        
        # If 8 chars (e.g. FFD9EAD3), strip first 2
        if len(hex_str) == 8:
            return hex_str[2:]
        return hex_str
    
    
    def load_all_questions(self):
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file missing at: {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found.")
            
        sheet = wb[self.sheet_name]
        
        headers = {}
        for cell in sheet[1]:
            if cell.value: headers[str(cell.value).strip()] = cell.column

        q_col = headers.get("Question")
        t_col = headers.get("Topics") # This is where the color lives

        if not q_col:
            raise ValueError("Could not find 'Question' column.")
        
        self.questions_cache = []
        for row in sheet.iter_rows(min_row=2, min_col=q_col, max_col=q_col):
            cell = row[0]
            if cell.value:
                # 1. Extract Link
                link = cell.hyperlink.target if cell.hyperlink else None
                
                # 2. Extract Topic & Color
                topic = "General"
                difficulty = "Medium" # Default fallback
                
                if t_col:
                    topic_cell = sheet.cell(row=cell.row, column=t_col)
                    topic = topic_cell.value
                    
                    # COLOR LOGIC
                    if topic_cell.fill and topic_cell.fill.fgColor.rgb:
                        raw_hex = topic_cell.fill.fgColor.rgb
                        clean_hex = self._normalize_hex(raw_hex)
                        
                        if clean_hex in COLOR_MAP:
                            difficulty = COLOR_MAP[clean_hex]

                self.questions_cache.append({
                    "title": str(cell.value).strip(),
                    "link": link,
                    "topic": str(topic).strip(),
                    "difficulty": difficulty,
                    "row": cell.row
                })
        print(f"✅ Loaded {len(self.questions_cache)} questions.")


    def get_new_question(self, completed_titles):
        if not self.questions_cache:
            self.load_all_questions()

        available = [q for q in self.questions_cache if q['title'] not in completed_titles]

        if not available:
            return None 

        return random.choice(available)